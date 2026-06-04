"""
============================================================================
GLMsingle Multi-Subject Pipeline
============================================================================
This script runs GLMsingle analysis for multiple subjects with per-subject
TR specification, comprehensive error checking, and detailed logging.

Run pipeline in terminal:
> conda activate env_ana_fMRI
> python conditioned_inhibition/analysis/fMRI/rsa_analysis/glmsingle_pipeline_withOutcomes.py

Save version:
> nohup python conditioned_inhibition/analysis/fMRI/rsa_analysis/glmsingle_pipeline_withOutcomes.py > pipeline_run.log 2>&1 &
> tail -f pipeline_run.log

Author: Sarah Salzgeber & Rishabh Garg
Date: May 2026
============================================================================
"""

# NOTE Development
# if processing_summary.csv exists and pipeline is rerun, even for single subject 
# the file is overwritten (old info about subject who already ran successfully or not)s

import numpy as np
import pandas as pd
import scipy
import scipy.stats as stats
import scipy.io as sio
import matplotlib.pyplot as plt
import nibabel as nib

import os
from os.path import join, exists, split
import sys
import time
import urllib.request
import copy
import warnings
from tqdm import tqdm
from pprint import pprint
warnings.filterwarnings('ignore')

import ast
import json
from pathlib import Path
from nilearn import image

import glmsingle
from glmsingle.glmsingle import GLM_single

import argparse
import contextlib

# ============================================================================
# CONFIGURATION
# ============================================================================

# ---- Change this one line to switch between datasets ----
DATASET = "full_sample"  # Options: "full_sample" | "pilot"

_LOCAL_BASE = Path("/Users/sasalz/switchdrive/Institution/causal_knowledge/conditioned_inhibition/results_local")

_CONFIGS = {
    "full_sample": dict(
        base_dir        = Path("/Volumes/2025/lor_salzgeber_hare_causal_knowledge/data/fmri_data"),
        fmriprep_subdir = Path("derivatives/no_sdc/fmriprep"),
        output_dir      = _LOCAL_BASE / "derivatives" / "GLMsingle" / "glmsingle_outputs_fullsample",
        task_name       = "causal",
        motion_qc_file  = Path("/Volumes/2025/lor_salzgeber_hare_causal_knowledge/data/fmri_data/motion_qc_fmriprep.xlsx"),
    ),
    "pilot": dict(
        base_dir        = Path("/Volumes/2025/lor_salzgeber_hare_causal_knowledge/data/pilot_data"),
        fmriprep_subdir = Path("derivatives/fmriprep25/no_fmap_correction"),
        output_dir      = _LOCAL_BASE / "derivatives" / "GLMsingle" / "glmsingle_outputs",
        task_name       = "cond",
        motion_qc_file  = None,  # no QC file for pilot; all subjects in fmriprep dir are used
    ),
}

_cfg         = _CONFIGS[DATASET]
BASE_DIR     = _cfg["base_dir"]
FMRIPREP_DIR = BASE_DIR / _cfg["fmriprep_subdir"]
OUTPUT_DIR   = _cfg["output_dir"]
TASK_NAME    = _cfg["task_name"]
MOTION_QC_FILE = _cfg["motion_qc_file"]

BIDS_DIR = BASE_DIR / "bids"

# EVENT duration (in seconds) - NOTE: can only specify one value for all events => same duration for cues and outcomes
SET_STIMDUR = 1.5

def load_tr_from_bids(bids_dir, subject, task_name):
    """Read TR from the first func JSON sidecar found in BIDS for this subject/task."""
    jsons = sorted((bids_dir / f"sub-{subject}" / "func").glob(
        f"sub-{subject}_task-{task_name}_run-*_bold.json"))
    if not jsons:
        raise FileNotFoundError(
            f"No func JSON sidecar found for sub-{subject} task-{task_name} in {bids_dir}"
        )
    with open(jsons[0]) as f:
        return json.load(f)["RepetitionTime"]


def load_excluded_subjects(qc_file):
    """Return set of subject IDs (e.g. '007') where excluded_final==1 in the QC spreadsheet."""
    import openpyxl
    wb = openpyxl.load_workbook(qc_file)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    subj_idx = headers.index("Subject")
    excl_idx = headers.index("excluded_final")
    return {
        row[subj_idx].replace("sub-", "")
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[subj_idx] and row[excl_idx] == 1
    }


def discover_subjects(fmriprep_dir, motion_qc_file=None):
    """Return sorted list of subject IDs found in fmriprep_dir, minus QC-excluded ones."""
    subs = sorted(
        p.name.replace("sub-", "")
        for p in fmriprep_dir.glob("sub-*") if p.is_dir()
    )
    if motion_qc_file is not None:
        excluded = load_excluded_subjects(motion_qc_file)
        print(f"  Excluding {len(excluded)} subjects per motion QC (excluded_final==1): {sorted(excluded)}")
        subs = [s for s in subs if s not in excluded]
    return subs


@contextlib.contextmanager
def subject_logger(log_path):
    """
    Context manager to write all prints for one subject to a log file,
    while still mirroring them to the console.
    """
    class Tee(object):
        def __init__(self, *files):
            self.files = files
        def write(self, obj):
            for f in self.files:
                f.write(obj)
                f.flush()
        def flush(self):
            for f in self.files:
                f.flush()

    original_stdout = sys.stdout
    original_stderr = sys.stderr
    with open(log_path, "w", encoding="utf-8") as log_f:
        tee_stdout = Tee(original_stdout, log_f)
        tee_stderr = Tee(original_stderr, log_f)
        sys.stdout = tee_stdout
        sys.stderr = tee_stderr
        try:
            yield
        finally:
            sys.stdout = original_stdout
            sys.stderr = original_stderr

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def clean_stimuli(val):
    """
    Convert stimulus values to clean underscore-separated strings.
    Handles lists, list-like strings, and regular strings.
    
    Parameters:
        val: stimulus value (list, string, or other)
    
    Returns:
        Clean string representation of stimulus
    """
    # If value is a list, join it
    if isinstance(val, list):
        return "_".join(val)
    # If value is a string and looks like a list, parse then join
    if isinstance(val, str):
        if val.startswith("[") and val.endswith("]"):
            try:
                val_list = ast.literal_eval(val)
                return "_".join(val_list)
            except Exception:
                pass
        # Otherwise, just use string as-is and remove spaces/quotes
        return val.replace("'", "").replace('"', '').replace(",", "_").replace(" ", "_")
    return str(val)

def split_outcome_column(events):
    """
    Split outcome column into outcome_reward and outcome_hidden (nuisance regressors).`
    Parameters:
        events (pd.DataFrame): Events dataframe
    Returns:
        None
    """
    # check if columns already exist to avoid overwriting
    if 'outcome_reward' in events.columns or 'outcome_hidden' in events.columns:
        return 
    events['outcome_reward'] = events['outcome'] != 'hidden'
    events['outcome_hidden'] = events['outcome'] == 'hidden'
    return 

def compute_stimulus_duration(events):
    """
    Compute average stimulus duration from events.
    
    Parameters:
        events (pd.DataFrame): Events dataframe
    
    Returns:
        float: Mean stimulus duration in seconds
    """
    events['stimdur'] = events['cue_end'] - events['cue_start']
    return 

def load_and_validate_events(events_file):
    """
    Load and validate events CSV file.
    
    Parameters:
        events_file (Path): Path to events CSV file
    
    Returns:
        pd.DataFrame: Cleaned and sorted events dataframe
    
    Raises:
        FileNotFoundError: If file does not exist
        ValueError: If required columns are missing
    """
    # Check file exists
    if not events_file.exists():
        raise FileNotFoundError(f"Events file not found: {events_file}")
    
    # Load CSV
    events = pd.read_csv(events_file)
    
    # Check required columns
    required_cols = ['run', 'stimuli', 'cue_start', 'cue_end', 'trial_start']
    missing_cols = [col for col in required_cols if col not in events.columns]
    if missing_cols:
        raise ValueError(f"Events file missing required columns: {missing_cols}")
    
    # Sort by run and trial timing for chronological order
    events = events.sort_values(['run', 'cue_start']).reset_index(drop=True)
    
    # Clean stimulus names
    events['stimuli_clean'] = events['stimuli'].apply(clean_stimuli)
    
    # Create unique trial labels: trial_000_stimulus_name
    events['trial_label'] = [
        f"trial_{i:03d}_{row.stimuli_clean}"
        for i, row in enumerate(events.itertuples(index=False), start=1)
    ]
    
    return events

def load_and_validate_bold_files(bold_dir, subject, task):
    """
    Load and validate preprocessed BOLD files from fMRIPrep.
    
    Parameters:
        bold_dir (Path): Directory containing BOLD files
        subject (str): Subject ID (e.g., "014")
        task (str): Task name (e.g., "cond")
    
    Returns:
        list: Sorted list of Path objects to BOLD files
    
    Raises:
        FileNotFoundError: If no BOLD files found
        ValueError: If BOLD files have inconsistent dimensions
    """
    # Find preprocessed BOLD files in standard fMRIPrep space
    bold_files = sorted(
        bold_dir.glob(
            f"sub-{subject}_task-{task}_run-*_space-MNI152NLin2009cAsym_desc-preproc_bold.nii.gz"
        )
    )
    
    if not bold_files:
        raise FileNotFoundError(
            f"No preprocessed BOLD files found for sub-{subject} task-{task}"
        )
    # Check consistency: all files should have same dimensions (except time)
    reference_img = image.load_img(str(bold_files[0]))
    reference_shape = reference_img.shape[:3]  # x, y, z dimensions

    for bold_file in bold_files[1:]:
        img = image.load_img(str(bold_file))
        if img.shape[:3] != reference_shape:
            raise ValueError(
                f"Inconsistent BOLD dimensions: {bold_file.name} has shape {img.shape}, "
                f"expected {(*reference_shape, '*')}"
            )
    
    return bold_files
  
def create_design_matrices(events, bold_dir, subject, task, tr):
    """
    Create design matrices for each run from events and BOLD files.
    
    Parameters:
        events (pd.DataFrame): Events dataframe with cue_start, cue_end, run, stimuli_clean
        bold_dir (Path): Directory containing BOLD files
        subject (str): Subject ID
        task (str): Task name
        tr (float): Repetition time in seconds
    
    Returns:
        tuple: (design_matrices list, fmri_data list, unique_stimuli list)
    
    Raises:
        ValueError: If events don't align with BOLD data or mismatch in runs
    """
    design_matrices = []
    fmri_data = []
    
    # Get all unique stimulus types across all runs
    unique_stimuli = sorted(events['stimuli_clean'].unique())
    
    # Create stimulus-to-column mapping
    stim2col = {s: i for i, s in enumerate(unique_stimuli)}

    # add outcome conditions as nuisance regressors
    NUISANCE_CONDITIONS = ['outcome_reward', 'outcome_hidden']  # or just ['outcome']
    all_conditions = unique_stimuli + NUISANCE_CONDITIONS
    nuisance2col = {n: len(unique_stimuli) + i for i, n in enumerate(NUISANCE_CONDITIONS)}
    reg2label = {**stim2col, **nuisance2col}
    
    # Load BOLD files and create corresponding design matrices
    bold_files = load_and_validate_bold_files(bold_dir, subject, task)
    
    print(f"\n  Creating design matrices for {len(bold_files)} runs...")
    
    for bold_path in bold_files:
        # Extract run number from filename (e.g., "run-01" -> 1)
        try:
            run_num = int(bold_path.name.split('run-')[1][:2])
        except (IndexError, ValueError):
            raise ValueError(f"Could not extract run number from {bold_path.name}")
        
        # Load BOLD data (shape: x, y, z, volumes)
        img = image.load_img(str(bold_path))
        data_4d = img.get_fdata()
        n_volumes = data_4d.shape[-1]
        
        # Get events for this run
        run_events = events[events['run'] == run_num].copy()
        if run_events.empty:
            print(f"    WARNING: No events found for run {run_num}, skipping...")
            continue
        
        # Use first event's timing as scan start reference
        scan_start = run_events['trial_start'].min()
        
        # Initialize design matrix: (n_volumes, n_conditions)
        dm = np.zeros((n_volumes, len(all_conditions)))
        
        # Fill design matrix: mark stimulus onsets at appropriate TR indices
        for _, row in run_events.iterrows():
            
            # Calculate TR index using floor division to assign to the TR it falls within
            onset_sec = row['cue_start'] - scan_start
            if onset_sec < 0:
                raise ValueError(f"Event onset CUE precedes scan start: {onset_sec:.4f}s")
            # floor division assigns the event onset to the TR during which the stimulus occurred, not the nearest TR boundary (rounding)
            onset_tr = int(onset_sec / tr)  
            
            # Ensure TR index is within valid range
            stim_col = stim2col[row['stimuli_clean']]     
            if 0 <= onset_tr < n_volumes:
                dm[onset_tr, stim_col] = 1.0
            else:
                print(f"    WARNING: Run {run_num}, trial at TR {onset_tr} outside range [0, {n_volumes})")
            
            # add outcome onsets 
            out_onset_sec = row['outcome_start'] - scan_start
            if out_onset_sec < 0:
                raise ValueError(f"Event onset OUT precedes scan start: {out_onset_sec:.4f}s")
            out_onset_tr = int(out_onset_sec / tr)    
            outcome_type = 'outcome_reward' if row['outcome'] != 'hidden' else 'outcome_hidden'
            out_col = nuisance2col[outcome_type]
            if 0 <= out_onset_tr < n_volumes:
                dm[out_onset_tr, out_col] = 1.0
            else:
                print(f"    WARNING: Run {run_num}, trial at TR {out_onset_tr} outside range [0, {n_volumes})")
        
        design_matrices.append(dm)
        fmri_data.append(data_4d)
        
        event_count = int(np.sum(dm > 0))
        print(f"    Run {run_num}: design matrix shape {dm.shape}, marked {event_count} events")
    
    if not design_matrices:
        raise ValueError(f"No valid design matrices created for sub-{subject}")
    
    return design_matrices, fmri_data, all_conditions, NUISANCE_CONDITIONS, reg2label

def verify_design_matrices(design_matrices, events, dm_cols, nuisance_cols):
    """
    Basic consistency check between events and design matrices.
    Prints, for each run, the expected vs. actual number of events per condition.
    """
    print("\n[Verification] Design matrix vs events per run:\n")

    # Count events per run and condition from events table
    event_counts_cues = (
        events.groupby(['run', 'stimuli_clean'])
              .size()
              .unstack(fill_value=0)
    )
    event_counts_out = (
        events.groupby(['run', 'outcome_reward', 'outcome_hidden'])
              .size()
    ).reset_index(name='count')

    for run_index, dm in enumerate(design_matrices, start=1):
        print(f"  Run {run_index}:")
        # Actual counts from design matrix (one event per '1' in column)
        dm_counts = dm.sum(axis=0)

        for col_idx, stim in enumerate(dm_cols):
            if stim in nuisance_cols: 
                event_counts_out_run = event_counts_out[event_counts_out['run'] == run_index]
                expected = event_counts_out_run[event_counts_out_run[stim] == True]['count'].values[0]
            else:
                expected = event_counts_cues.loc[run_index, stim] if stim in event_counts_cues.columns and run_index in event_counts_cues.index else 0
            actual = int(dm_counts[col_idx])
            status = "OK" if expected == actual else "MISMATCH"
            print(f"    {stim:30s} expected={expected:3d}  actual={actual:3d}  [{status}]")
    print("")

def configure_glmsingle_options():
    """
    Configure GLMsingle options dictionary.
    
    Returns:
        dict: GLMsingle configuration options
    """
    opt = dict()
    
    # Session indicator (all runs from same session)
    # Will be updated with actual number of runs per subject
    opt['sessionindicator'] = None  # Set during processing
    
    # Enable all GLMsingle features
    opt['wantlibrary'] = 1       # Use HRF library optimization
    opt['wantglmdenoise'] = 1    # Use GLMdenoise (noise regressors)
    opt['wantfracridge'] = 1     # Use fractional ridge regression
    
    # Save only final outputs (betasmd - denoised + ridge betas)
    # [old betas, R2, HRF, betasmd]
    opt['wantfileoutputs'] = [1, 1, 1, 1]
    
    return opt

def run_glmsingle(design_matrices, fmri_data, stimdur, tr, subject_output, figures_dir):
    """
    Run GLMsingle fitting on design matrices and fMRI data.
    
    Parameters:
        design_matrices (list): List of design matrices (n_volumes, n_conditions)
        fmri_data (list): List of fMRI data arrays (x, y, z, n_volumes)
        stimdur (float): Stimulus duration in seconds
        tr (float): Repetition time in seconds
        subject_output (Path): Output directory for subject results
        figures_dir (Path): Directory for diagnostic figures
    
    Returns:
        dict: GLMsingle results dictionary
    
    Raises:
        RuntimeError: If GLMsingle fitting fails
    """
    # Create GLMsingle configuration
    opt = configure_glmsingle_options()
    opt['sessionindicator'] = np.ones((1, len(design_matrices)), dtype=int)
    
    print(f"  GLMsingle configuration:")
    print(f"    - HRF library optimization: {bool(opt['wantlibrary'])}")
    print(f"    - GLMdenoise: {bool(opt['wantglmdenoise'])}")
    print(f"    - Fractional ridge: {bool(opt['wantfracridge'])}")
    print(f"    - Stimulus duration: {stimdur:.2f}s")
    print(f"    - TR: {tr:.3f}s")
    print(f"    - Number of runs: {len(design_matrices)}\n")
    
    # Create GLMsingle object and fit
    try:
        glmsingle_obj = GLM_single(opt)
        print(f"  Running GLMsingle (this may take several minutes)...\n")
        results = glmsingle_obj.fit(
            design_matrices,
            fmri_data,
            stimdur,
            tr,
            outputdir=str(subject_output),
            figuredir=str(figures_dir)
        )
        print(f"\n  ✓ GLMsingle fitting complete!")
        return results
    except Exception as e:
        raise RuntimeError(f"GLMsingle fitting failed: {str(e)}")

def extract_and_save_outputs(subject, subject_output, bold_dir, reg2label, task_name):
    """
    Extract GLMsingle outputs and save as NIfTI and CSV files.
    
    Parameters:
        subject (str): Subject ID
        subject_output (Path): Output directory for subject
        bold_dir (Path): Directory with reference BOLD file for affine/header
        reg2label (dict): Mapping of design matrix column indices to condition labels
    
    Raises:
        FileNotFoundError: If GLMsingle output files not found
        ValueError: If output shapes don't match reference BOLD
    """
    # Find and load the GLMsingle typed results file
    typed_file = subject_output / "TYPED_FITHRF_GLMDENOISE_RR.npy"
    if not typed_file.exists():
        raise FileNotFoundError(f"GLMsingle results file not found: {typed_file}")
    
    typed = np.load(str(typed_file), allow_pickle=True).item()
    
    # Extract betas and R² maps
    betas = typed['betasmd']
    R2 = typed['R2']
    hrf_index = typed['HRFindex']
    
    print(f"\n  Extracting and saving outputs:")
    print(f"    - Betas shape: {betas.shape}")
    print(f"    - R² shape: {R2.shape}")
    print(f"    - HRF index shape: {hrf_index.shape}")
    
    # load reference BOLD file for affine matrix and header info
    bold_files = sorted(bold_dir.glob(f"sub-{subject}_task-*_run-01_space-MNI152NLin2009cAsym_desc-preproc_bold.nii.gz"))
    if not bold_files:
        raise FileNotFoundError(f"Reference BOLD file not found for sub-{subject}")
    img0 = image.load_img(str(bold_files[0]))
    
    #------- Convert betas to NIfTI image and save
    betas_img = image.new_img_like(img0, betas)
    betas_nifti = subject_output / f"sub-{subject}_task-{task_name}_space-MNI152NLin2009cAsym_desc-glmSingle_betas.nii.gz"
    betas_img.to_filename(str(betas_nifti))
    print(f"    ✓ Saved betas: {betas_nifti.name}")
    
    #------- Convert R² to NIfTI image and save
    R2_img = image.new_img_like(img0, R2)
    R2_nifti = subject_output / f"sub-{subject}_glmSingle_R2.nii.gz"
    R2_img.to_filename(str(R2_nifti))
    print(f"    ✓ Saved R²: {R2_nifti.name}")

    #------- save R2 as npy file as well 
    R2_npy = subject_output / f"sub-{subject}_R2.npy"
    np.save(R2_npy, R2)
    print(f"    ✓ Saved R² (NumPy): {R2_npy.name}")
    if 'R2run' in typed:
        R2run = typed['R2run']
        R2run_npy = subject_output / f"sub-{subject}_R2run.npy"
        np.save(R2run_npy, R2run)
        print(f"    ✓ Saved run-wise R² (NumPy): {R2run_npy.name}")

    #------- Convert HRF index to NIfTI image and save
    hrf_img = image.new_img_like(img0, hrf_index)
    hrf_nifti = subject_output / f"sub-{subject}_glmSingle_HRFindex.nii.gz"
    hrf_img.to_filename(str(hrf_nifti))
    print(f"    ✓ Saved HRF index map: {hrf_nifti.name}")

    #------- Plot and save histogram of HRF index distribution across voxels
    plt.figure()
    plt.hist(hrf_index.flatten(), bins=np.unique(hrf_index).size)
    plt.xlabel("HRF index")
    plt.ylabel("Voxel count")
    plt.title(f"HRF distribution (sub-{subject})")
    hrf_hist_file = subject_output / "figures" / f"sub-{subject}_HRFindex_hist.png"
    plt.savefig(hrf_hist_file)
    plt.close()
    print(f"    ✓ Saved HRF histogram: {hrf_hist_file.name}")
    
    #------- Load and save trial information (stimulus order)
    designinfo_file = subject_output / "DESIGNINFO.npy"
    if not designinfo_file.exists():
        raise FileNotFoundError(f"Design info file not found: {designinfo_file}")
    
    designinfo = np.load(str(designinfo_file), allow_pickle=True).item()
    stimorder = designinfo['stimorder']
    
    #------- Create beta to trial label mapping as CSV file 
    label2reg = {v: k for k, v in reg2label.items()}
    df_trialinfo = pd.DataFrame({
        'trial_id': np.arange(len(stimorder)),
        'regressor': stimorder,
        'reg_label': pd.Series(stimorder).map(label2reg)
    })
    trialinfo_csv = subject_output / f"sub-{subject}_glmSingle_betas_info.csv"
    df_trialinfo.to_csv(str(trialinfo_csv), index=False)
    print(f"    ✓ Saved trial info (CSV): {trialinfo_csv.name}")

    #------- create nifty with cue betas only 
    cue_mask = ~df_trialinfo['reg_label'].str.contains('outcome', na=False)
    cue_indices = df_trialinfo.loc[cue_mask, 'trial_id'].values
    betas_cues = betas[..., cue_indices]
    betas_cues_img = image.new_img_like(img0, betas_cues)
    betas_cues_nifti = subject_output / f"sub-{subject}_task-{task_name}_space-MNI152NLin2009cAsym_desc-glmSingle_betas_CUES.nii.gz"
    betas_cues_img.to_filename(str(betas_cues_nifti))
    print(f"    ✓ Saved cue betas: {betas_cues_nifti.name}")

    #------- save trial info for cues only 
    df_trialinfo_cues = df_trialinfo[cue_mask].reset_index(drop=True)
    trialinfo_cues_csv = subject_output / f"sub-{subject}_glmSingle_betas_CUES_info.csv"
    df_trialinfo_cues.to_csv(str(trialinfo_cues_csv), index=False)
    print(f"    ✓ Saved trial info for cues (CSV): {trialinfo_cues_csv.name}")

    #------- Compute and save mean and std betas
    beta_mean = np.mean(betas, axis=-1)
    beta_mean_img = image.new_img_like(img0, beta_mean)
    beta_mean_nifti = subject_output / f"sub-{subject}_task-{task_name}_space-MNI152NLin2009cAsym_desc-glmSingle_betas_AVG.nii.gz"
    beta_mean_img.to_filename(str(beta_mean_nifti))
    print(f"    ✓ Saved cue betas: {beta_mean_nifti.name}")
    
    beta_std = np.std(betas, axis=-1)
    beta_std_img = image.new_img_like(img0, beta_std)
    beta_std_nifti = subject_output / f"sub-{subject}_task-{task_name}_space-MNI152NLin2009cAsym_desc-glmSingle_betas_STD.nii.gz"
    beta_std_img.to_filename(str(beta_std_nifti))
    print(f"    ✓ Saved cue betas: {beta_std_nifti.name}")

    beta_cues_mean = np.mean(betas_cues, axis=-1)
    beta_cues_mean_img = image.new_img_like(img0, beta_cues_mean)
    beta_cues_mean_nifti = subject_output / f"sub-{subject}_task-{task_name}_space-MNI152NLin2009cAsym_desc-glmSingle_betas_CUES_AVG.nii.gz"
    beta_cues_mean_img.to_filename(str(beta_cues_mean_nifti))
    print(f"    ✓ Saved cue betas: {beta_cues_mean_nifti.name}")

    beta_cues_std = np.std(betas_cues, axis=-1)
    beta_cues_std_img = image.new_img_like(img0, beta_cues_std)
    beta_cues_std_nifti = subject_output / f"sub-{subject}_task-{task_name}_space-MNI152NLin2009cAsym_desc-glmSingle_betas_CUES_STD.nii.gz"
    beta_cues_std_img.to_filename(str(beta_cues_std_nifti))
    print(f"    ✓ Saved cue betas: {beta_cues_std_nifti.name}")

    # Compute tsnr and save as NIfTI
    tsnr = beta_mean / (beta_std + 1e-6)
    tsnr_img = image.new_img_like(img0, tsnr)
    tsnr_nifti = subject_output / f"sub-{subject}_task-{task_name}_space-MNI152NLin2009cAsym_desc-glmSingle_betas_TSNR.nii.gz"
    tsnr_img.to_filename(str(tsnr_nifti))
    print(f"    ✓ Saved cue betas: {tsnr_nifti.name}")

    # Compute tsnr and save as NIfTI
    tsnr_cues = beta_cues_mean / (beta_cues_std + 1e-6)
    tsnr_cues_img = image.new_img_like(img0, tsnr_cues)
    tsnr_cues_nifti = subject_output / f"sub-{subject}_task-{task_name}_space-MNI152NLin2009cAsym_desc-glmSingle_betas_CUES_TSNR.nii.gz"
    tsnr_cues_img.to_filename(str(tsnr_cues_nifti))
    print(f"    ✓ Saved cue betas: {tsnr_cues_nifti.name}")

# ============================================================================
# SINGLE SUBJECT PIPELINE
# ============================================================================

def process_single_subject(subject, tr, bids_dir, fmriprep_dir, output_dir, task_name, set_stimdur):
    """
    Complete GLMsingle pipeline for a single subject.
    
    Parameters:
        subject (str): Subject ID (e.g., "014")
        tr (float): Repetition time in seconds
        bids_dir (Path): BIDS directory
        fmriprep_dir (Path): fMRIPrep derivatives directory
        output_dir (Path): Output directory for GLMsingle results
        task_name (str): Task name
        set_stimdur (float or None): Set stimulus duration
    
    Returns:
        bool: True if successful, False if failed
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING SUBJECT: sub-{subject}")
    print(f"{'='*70}")

    try:
        # ====== STEP 1: Load and validate events ======
        print(f"\n[1/6] Loading events...")
        events_dir = bids_dir / f"sub-{subject}" / "events"
        if not events_dir.exists():
            raise FileNotFoundError(f"Events directory not found: {events_dir}")
        
        # Find events file (usually a CSV with timestamp in name)
        events_files = list(events_dir.glob("*phase2*.csv"))
        if not events_files:
            raise FileNotFoundError(
                f"No phase2 events CSV files found in {events_dir} "
                f"(expected something like data_sub-{subject}_phase2-*.csv)"
            )
        
        events_file = events_files[0]  # Use first CSV found
        print(f"  ✓ Found events file - !!! PLEASE CHECK !!! - : {events_file.name}")
        
        events = load_and_validate_events(events_file)
        print(f"  ✓ Loaded {len(events)} trials")
        print(f"  ✓ Conditions: {list(events['stimuli_clean'].unique())}")
        print(f"  ✓ Runs: {sorted(events['run'].unique())}")
        
        # ====== STEP 2: Load and validate BOLD files ======
        print(f"\n[2/6] Loading preprocessed BOLD files...")
        bold_dir = fmriprep_dir / f"sub-{subject}" / "func"
        if not bold_dir.exists():
            raise FileNotFoundError(f"BOLD directory not found: {bold_dir}")
        
        bold_files = load_and_validate_bold_files(bold_dir, subject, task_name)
        print(f"  ✓ Found {len(bold_files)} BOLD files")
        for bf in bold_files:
            img = image.load_img(str(bf))
            print(f"    - {bf.name}: shape {img.shape}")
        
        # ====== STEP 3: Compute stimulus duration and split outcome ======
        print(f"\n[3/6] Compare stimulus duration...")
        split_outcome_column(events)
        compute_stimulus_duration(events)
        print(f"  ! CHECK: Specified stimulus duration {set_stimdur:.2f}s")
        print(f"         - Subject's mean stimulus duration: {events['stimdur'].mean():.2f}s")
        print(f"            - Range: {events['stimdur'].min():.2f}s to {events['stimdur'].max():.2f}s")
        print(f"         - Subject's mean RT (sec): {events['cue_rt_sec'].mean():.2f}s")
        print(f"            - Range: {events['cue_rt_sec'].min():.2f}s to {events['cue_rt_sec'].max():.2f}s")
        
        # ====== STEP 4: Create design matrices ======
        print(f"\n[4/6] Creating design matrices...")
        design_matrices, fmri_data, cond_cols, nuisance_cols, reg2label = create_design_matrices(
            events, bold_dir, subject, task_name, tr
        )
        print(f"  ✓ Created {len(design_matrices)} design matrices")
        print(f"  ✓ Conditions modeled: {cond_cols}")
        verify_design_matrices(design_matrices, events, cond_cols, nuisance_cols)
        
        # ====== STEP 5: Run GLMsingle ======
        print(f"\n[5/6] Running GLMsingle...")
        subject_output = output_dir / f"sub-{subject}"
        figures_dir = subject_output / "figures"
        subject_output.mkdir(parents=True, exist_ok=True)
        figures_dir.mkdir(parents=True, exist_ok=True)
        
        run_glmsingle(design_matrices, fmri_data, set_stimdur, tr, subject_output, figures_dir)
        
        # ====== STEP 6: Extract and save outputs ======
        print(f"\n[6/6] Extracting and saving outputs...")
        extract_and_save_outputs(subject, subject_output, bold_dir, reg2label, task_name)
        
        print(f"\n{'='*70}")
        print(f"✓ SUBJECT sub-{subject} COMPLETED SUCCESSFULLY")
        print(f"{'='*70}\n")
        
        return True
        
    except Exception as e:
        print(f"\n{'='*70}")
        print(f"✗ ERROR PROCESSING SUBJECT sub-{subject}")
        print(f"{'='*70}")
        print(f"Error: {str(e)}\n")
        return False

# ============================================================================
# MAIN EXECUTION
# ============================================================================
if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="GLMsingle multi-subject pipeline")
    parser.add_argument(
        "--subject", type=str, default=None,
        help="Run only one subject, e.g. --subject sub-064 or --subject 064"
    )
    args = parser.parse_args()

    print(f"\n{'='*70}")
    print("GLMsingle Multi-Subject Pipeline")
    print(f"{'='*70}\n")

    # Validate base directories exist
    print("Validating directories...")
    for dir_path, dir_name in [
        (BIDS_DIR, "BIDS"),
        (FMRIPREP_DIR, "fMRIPrep derivatives"),
    ]:
        if not dir_path.exists():
            print(f"✗ {dir_name} directory not found: {dir_path}")
            sys.exit(1)
        print(f"  ✓ {dir_name}: {dir_path}")

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"  ✓ Output directory: {OUTPUT_DIR}\n")

    # Track results
    results_summary = []

    # ========== DISCOVER / SELECT SUBJECTS ==========
    if args.subject:
        # Accept both "sub-064" and "064"
        sub_id = args.subject.replace("sub-", "")
        subjects = [sub_id]
        print(f"Single-subject mode: sub-{sub_id}\n")
    else:
        print(f"Discovering subjects for dataset '{DATASET}'...")
        subjects = discover_subjects(FMRIPREP_DIR, MOTION_QC_FILE)
        print(f"  {len(subjects)} subjects to process: {subjects}\n")

    # ========== LOOP OVER ALL SUBJECTS ==========
    print(f"Processing {len(subjects)} subjects:\n")
    for subject in subjects:
        tr = load_tr_from_bids(BIDS_DIR, subject, TASK_NAME)

        # Check completion before opening the log — skipped subjects must not
        # overwrite their existing log file.
        completion_marker = (
            OUTPUT_DIR / f"sub-{subject}"
            / f"sub-{subject}_task-{TASK_NAME}_space-MNI152NLin2009cAsym_desc-glmSingle_betas_CUES_TSNR.nii.gz"
        )
        if completion_marker.exists():
            print(f"\n{'='*70}")
            print(f"SKIPPING sub-{subject} — already completed (found {completion_marker.name})")
            print(f"{'='*70}\n")
            results_summary.append({'subject': subject, 'tr': tr, 'status': 'SKIPPED'})
            continue

        subject_log = OUTPUT_DIR / f"sub-{subject}_log.txt"
        with subject_logger(subject_log):
            success = process_single_subject(
                subject=subject,
                tr=tr,
                bids_dir=BIDS_DIR,
                fmriprep_dir=FMRIPREP_DIR,
                output_dir=OUTPUT_DIR,
                task_name=TASK_NAME,
                set_stimdur=SET_STIMDUR
            )

        results_summary.append({
            'subject': subject,
            'tr': tr,
            'status': 'SUCCESS' if success else 'FAILED'
        })
    
    # ========== PRINT SUMMARY ==========
    print(f"\n{'='*70}")
    print("PIPELINE SUMMARY")
    print(f"{'='*70}\n")
    
    df_summary = pd.DataFrame(results_summary)
    print(df_summary.to_string(index=False))
    
    n_success = (df_summary['status'] == 'SUCCESS').sum()
    n_total = len(df_summary)
    print(f"\n✓ Completed: {n_success}/{n_total} subjects\n")
    
    # Save summary
    summary_file = OUTPUT_DIR / "processing_summary.csv"
    df_summary.to_csv(str(summary_file), index=False)
    print(f"Summary saved to: {summary_file}\n")